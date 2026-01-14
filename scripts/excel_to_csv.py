from openpyxl import load_workbook
import csv

input_file = "data/raw/NewsSumm_Dataset.xlsx"
output_file = "data/processed/newssumm_raw.csv"

# Open workbook in read-only mode
wb = load_workbook(filename=input_file, read_only=True)
ws = wb.active  # first sheet

with open(output_file, "w", newline="", encoding="utf-8") as f:
    writer = csv.writer(f)

    for i, row in enumerate(ws.iter_rows(values_only=True)):
        writer.writerow(row)
        if i % 50000 == 0:
            print(f"Written {i} rows")

wb.close()
print("Conversion completed successfully.")
